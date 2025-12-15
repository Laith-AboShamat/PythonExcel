#!/usr/bin/env python3
"""Automates daily Excel updates across BAL and sales order shortage trackers.

The script performs the following operations in sequence:
1) Builds BAL Sheet2 with Sheet1 columns (Item number through Napco weight (Kg)) filtered to CW available physical > 0 and copies the result into the inventory sheet of the raw shortages workbook.
2) Rebuilds "Sheet 1" from the sales order shortages workbook using the requested FILTER logic
   and writes the result into the Shortages+AllOreders sheet.
3) Filters BAL rows by warehouse code (STORE-002/010/027/041) and pushes the subset into
   the inv sheet of the color transfer workbook.
4) Rebuilds "Sheet 2" from the sales order shortages workbook (FILTER + CHOOSECOLS) and
   writes it into the perfect order workbook (new.shortages) and the CLR workbook (sheet 1).
5) Refreshes the perfect order inv sheet with the warehouse-filtered BAL rows.

You can either let the script auto-discover workbooks using the glob patterns below or provide
explicit file paths via command-line arguments, e.g.:

    python process_excel.py --bal-file "D:/Reports/BAL-01.12.xlsx" \
        --sales-file "D:/Reports/Sales order shortages-01.12.xlsx" \
        --inventory-file "D:/Reports/نواقص الخام لطلبيات الالوانPPP.xlsx" \
        --shortages-file "D:/Reports/نواقص الوان الخام PPP.xlsx" \
        --transfer-file "D:/Reports/نواقص الوان تحويل.xlsx" \
        --perfect-order-file "D:/Reports/perfect order start.xlsx" \
        --clr-file "D:/Reports/Sales order shortagesCLRs.xlsx"

Run the script from the folder that contains the Excel files (default: script directory) if you
prefer pattern discovery. Adjust the filename patterns below if your files use different names.

Double-clicking the script (or passing --gui) launches a window where you can browse and run the
process without using the command line. Package with PyInstaller to hand off a standalone .exe.
"""

import argparse
import json
import logging
import math
import sys
import threading
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo, TableList

BAL_PATTERN_DEFAULT = "BAL-*.xlsx"
SALES_PATTERN_DEFAULT = "Sales order shortages-*.xlsx"
SALES_SOURCE_SHEET = "salesorderremainingquantitiesRe"
SALES_HEADER_ROW_INDEX = 10  # zero-based index -> row 11 in Excel terms
SALES_COL_RANGE = ("B", "AQ")
FILTER_FALLBACK_MESSAGE = "Contact Fx. Provider"
WAREHOUSE_ALLOWED_VALUES = {"STORE-002", "STORE-010", "STORE-027", "STORE-041"}
WAREHOUSE_HEADER_KEYWORDS = ("warehouse", "مستودع")
FILTER1_EXCLUDE_Q = {
    "صيانة",
    "تبرعات",
    "عينات اكسسوارات",
    "عينات بروفيلات ألمنيوم",
    "طلبيات دائرة البحث و التطوير",
}
FILTER2_EXCLUDE_R = {
    "المنيوم لون فاتورة نابكو NF441",
    "المنيوم لون NF44",
    "اطوال ابيض",
    "المنيوم ابيض - مطابخ",
    "المنيوم ابيض",
    "NB101 ميشي مخرمش",
    "NB100 فضي طبيعي مخرمش",
    "MF000000",
    "اطوال خام",
}

SALES_COL_START_INDEX = column_index_from_string(SALES_COL_RANGE[0])
SALES_COL_END_INDEX = column_index_from_string(SALES_COL_RANGE[1])
ALL_SALES_LETTERS = [
    get_column_letter(idx)
    for idx in range(SALES_COL_START_INDEX, SALES_COL_END_INDEX + 1)
]
# Sheet1 mirrors the original FILTER output, so keep the full B:AQ span.
SHEET1_OUTPUT_LETTERS = list(ALL_SALES_LETTERS)
CHOOSECOLS_INDEXES = [1, 12, 13, 15, 16, 17, 30, 31]
SHEET2_COLUMN_LETTERS = [
    get_column_letter(SALES_COL_START_INDEX + offset - 1)
    for offset in CHOOSECOLS_INDEXES
]
SHEET1_REQUIRED_COLUMNS = {"B", "Q", "AE"}
SHEET2_REQUIRED_COLUMNS = {"B", "Q", "R", "AE"}
SHEET1_OUTPUT_REQUIRED_COLUMNS = set(SHEET1_OUTPUT_LETTERS)
SHEET2_OUTPUT_REQUIRED_COLUMNS = set(SHEET2_COLUMN_LETTERS)

WORKBOOK_PATTERNS = {
    "inventory": ["نواقص الخام لطلبيات الالوانPPP*.xlsx", "نواقص الوان الخام PPP*.xlsx"],
    "shortages_all_orders": ["نواقص الوان الخام PPP*.xlsx", "نواقص الخام لطلبيات الالوانPPP*.xlsx"],
    "transfer": ["نواقص الوان تحويل*.xlsx"],
    "perfect_order": ["perfect order start*.xlsx"],
    "clr": ["Sales order shortagesCLRs*.xlsx"],
}

BAL_REQUIRED_HEADERS = [
    "Item number",
    "Search name",
    "Warehouse",
    "Name",
    "CW physical inventory",
    "Physical inventory",
    "CW available physical",
    "Napco weight (Kg)",
]
BAL_AVAILABLE_HEADER = "CW available physical"
BAL_NULLISH_TEXT = {"", "null", "none"}


def default_root() -> Path:
    if getattr(sys, "frozen", False):  # PyInstaller bundle
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


CONFIG_PATH = Path.home() / ".process_excel_paths.json"


@dataclass
class WorkbookBundle:
    bal: Path
    sales: Path
    inventory: Path
    shortages: Path
    transfer: Path
    perfect_order: Path
    clr: Path


@dataclass
class SheetUpdate:
    sheet_name: Optional[str]
    rows: List[List[Any]]
    preserve_header: bool = False
    create_if_missing: bool = False


@dataclass
class SheetMetrics:
    workbook: Path
    sheet: str
    rows_written: int
    columns_written: int
    dry_run: bool


@dataclass
class FilterSnapshot:
    name: str
    input_rows: int
    output_rows: int


@dataclass
class PipelineReport:
    root: Path
    dry_run: bool
    duration_seconds: float
    filters: List[FilterSnapshot]
    sheets: List[SheetMetrics]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Refresh Excel trackers using BAL and sales shortage files.")
    parser.add_argument(
        "--root",
        type=Path,
        default=default_root(),
        help="Folder that contains the Excel files (default: script directory).",
    )
    parser.add_argument(
        "--bal-pattern",
        default=BAL_PATTERN_DEFAULT,
        help="Glob pattern for the BAL workbook (default: %(default)s).",
    )
    parser.add_argument(
        "--sales-pattern",
        default=SALES_PATTERN_DEFAULT,
        help="Glob pattern for the sales order shortages workbook (default: %(default)s).",
    )
    parser.add_argument(
        "--warehouse-column",
        default=None,
        help="Optional Excel column letter to force as the warehouse column in BAL (example: F).",
    )
    parser.add_argument("--bal-file", type=Path, default=None, help="Explicit path to the BAL workbook.")
    parser.add_argument(
        "--sales-file",
        type=Path,
        default=None,
        help="Explicit path to the Sales order shortages workbook.",
    )
    parser.add_argument(
        "--inventory-file",
        type=Path,
        default=None,
        help="Explicit path to the raw shortages workbook (inventory sheet update).",
    )
    parser.add_argument(
        "--shortages-file",
        type=Path,
        default=None,
        help="Explicit path to the PPP shortages workbook (Shortages+AllOreders sheet).",
    )
    parser.add_argument(
        "--transfer-file",
        type=Path,
        default=None,
        help="Explicit path to the color transfer workbook (inv sheet).",
    )
    parser.add_argument(
        "--perfect-order-file",
        type=Path,
        default=None,
        help="Explicit path to the perfect order workbook (new.shortages & inv sheets).",
    )
    parser.add_argument(
        "--clr-file",
        type=Path,
        default=None,
        help="Explicit path to the CLR workbook (Sheet1 update).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Load and validate everything without writing changes to disk.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Enable debug logging.",
    )
    parser.add_argument(
        "--gui",
        action="store_true",
        help="Launch the graphical interface.",
    )
    parser.add_argument(
        "--cli",
        action="store_true",
        help="Force command-line mode even without extra parameters.",
    )
    return parser.parse_args()


def setup_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format="%(levelname)s: %(message)s")


def locate_latest(pattern: str, root: Path) -> Path:
    matches = sorted(root.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    if not matches:
        raise FileNotFoundError(f"No files match pattern '{pattern}' under {root}")
    logging.debug("Selected %s for pattern %s", matches[0].name, pattern)
    return matches[0]


def locate_workbook(patterns: Sequence[str], root: Path) -> Path:
    for pattern in patterns:
        matches = sorted(root.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
        if matches:
            logging.debug("Selected %s for pattern %s", matches[0].name, pattern)
            return matches[0]
    raise FileNotFoundError(f"Could not locate any workbook matching {patterns} under {root}")


def resolve_explicit_path(path: Optional[Path], root: Path) -> Optional[Path]:
    if path is None:
        return None
    resolved = path if path.is_absolute() else (root / path).resolve()
    if not resolved.exists():
        raise FileNotFoundError(f"Workbook not found: {resolved}")
    logging.debug("Using explicit workbook path: %s", resolved)
    return resolved


def resolve_workbook(
    explicit: Optional[Path],
    patterns: Sequence[str],
    root: Path,
) -> Path:
    resolved = resolve_explicit_path(explicit, root)
    if resolved:
        return resolved
    return locate_workbook(patterns, root)


def resolve_latest_with_override(
    explicit: Optional[Path],
    pattern: str,
    root: Path,
) -> Path:
    resolved = resolve_explicit_path(explicit, root)
    if resolved:
        return resolved
    return locate_latest(pattern, root)


def load_preferences() -> Dict[str, str]:
    if not CONFIG_PATH.exists():
        return {}
    try:
        with CONFIG_PATH.open("r", encoding="utf-8") as handle:
            data = json.load(handle)
        if isinstance(data, dict):
            return {str(k): str(v) for k, v in data.items()}
    except Exception:
        logging.warning("Could not read preferences file '%s'.", CONFIG_PATH)
    return {}


def save_preferences(values: Dict[str, str]) -> None:
    try:
        CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
        with CONFIG_PATH.open("w", encoding="utf-8") as handle:
            json.dump(values, handle, ensure_ascii=False, indent=2)
    except Exception:
        logging.warning("Could not write preferences file '%s'.", CONFIG_PATH)


def resolve_sales_sheet_name(workbook_path: Path, target_name: str) -> str:
    with pd.ExcelFile(workbook_path) as xls:
        sheet_names = xls.sheet_names
    if target_name in sheet_names:
        return target_name
    normalized_target = target_name.strip().casefold()
    for name in sheet_names:
        if name.strip().casefold() == normalized_target:
            logging.info("Using sheet '%s' (matched requested '%s')", name, target_name)
            return name
    raise KeyError(
        f"Worksheet named '{target_name}' not found. Available sheets: {', '.join(sheet_names)}"
    )


def collect_workbook_bundle(args: argparse.Namespace, root: Path) -> WorkbookBundle:
    bal_path = resolve_latest_with_override(args.bal_file, args.bal_pattern, root)
    sales_path = resolve_latest_with_override(args.sales_file, args.sales_pattern, root)
    inventory_path = resolve_workbook(args.inventory_file, WORKBOOK_PATTERNS["inventory"], root)
    shortages_path = resolve_workbook(args.shortages_file, WORKBOOK_PATTERNS["shortages_all_orders"], root)
    transfer_path = resolve_workbook(args.transfer_file, WORKBOOK_PATTERNS["transfer"], root)
    perfect_order_path = resolve_workbook(args.perfect_order_file, WORKBOOK_PATTERNS["perfect_order"], root)
    clr_path = resolve_workbook(args.clr_file, WORKBOOK_PATTERNS["clr"], root)
    return WorkbookBundle(
        bal=bal_path,
        sales=sales_path,
        inventory=inventory_path,
        shortages=shortages_path,
        transfer=transfer_path,
        perfect_order=perfect_order_path,
        clr=clr_path,
    )


def auto_detect_paths(root: Path) -> Dict[str, Path]:
    detected: Dict[str, Path] = {}
    detected["bal"] = locate_latest(BAL_PATTERN_DEFAULT, root)
    detected["sales"] = locate_latest(SALES_PATTERN_DEFAULT, root)
    detected["inventory"] = locate_workbook(WORKBOOK_PATTERNS["inventory"], root)
    detected["shortages"] = locate_workbook(WORKBOOK_PATTERNS["shortages_all_orders"], root)
    detected["transfer"] = locate_workbook(WORKBOOK_PATTERNS["transfer"], root)
    detected["perfect_order"] = locate_workbook(WORKBOOK_PATTERNS["perfect_order"], root)
    detected["clr"] = locate_workbook(WORKBOOK_PATTERNS["clr"], root)
    return detected


def execute_pipeline(
    bundle: WorkbookBundle,
    warehouse_column: Optional[str],
    dry_run: bool,
    base_root: Path,
) -> PipelineReport:
    start_time = time.perf_counter()
    sheet_metrics: List[SheetMetrics] = []
    bal_raw_rows = load_bal_rows(bundle.bal)
    bal_selected_rows = select_bal_columns(bal_raw_rows, BAL_REQUIRED_HEADERS)
    bal_filtered_rows = filter_bal_rows_by_available(bal_selected_rows, BAL_AVAILABLE_HEADER)
    logging.info(
        "BAL availability filter retained %d of %d row(s)",
        max(len(bal_filtered_rows) - 1, 0),
        max(len(bal_selected_rows) - 1, 0),
    )
    warehouse_col_idx = detect_warehouse_column_index(bal_filtered_rows, warehouse_column)
    warehouse_rows = filter_bal_by_warehouse(bal_filtered_rows, warehouse_col_idx, WAREHOUSE_ALLOWED_VALUES)

    sales_df, header_map = load_sales_dataframe(bundle.sales)
    sheet1_df = build_sales_sheet_one(sales_df)
    sheet2_df = build_sales_sheet_two(sales_df)

    sheet1_rows = build_rows(
        sheet1_df,
        SHEET1_OUTPUT_LETTERS,
        header_map,
        FILTER_FALLBACK_MESSAGE,
        fill_value=0,
    )
    sheet2_rows = build_rows(
        sheet2_df,
        SHEET2_COLUMN_LETTERS,
        header_map,
        FILTER_FALLBACK_MESSAGE,
        fill_value=0,
    )

    sheet_metrics.extend(
        update_sales_workbook(bundle.sales, sheet1_rows, sheet2_rows, dry_run)
    )

    sheet_metrics.extend(
        apply_workbook_updates(
            bundle.bal,
            updates=[SheetUpdate("Sheet2", bal_filtered_rows, create_if_missing=True)],
            dry_run=dry_run,
        )
    )

    sheet_metrics.extend(
        apply_workbook_updates(
            bundle.inventory,
            updates=[SheetUpdate("inventory", bal_filtered_rows)],
            dry_run=dry_run,
        )
    )
    sheet_metrics.extend(
        apply_workbook_updates(
            bundle.shortages,
            updates=[SheetUpdate("Shortages+AllOreders", sheet1_rows, preserve_header=True)],
            dry_run=dry_run,
        )
    )
    sheet_metrics.extend(
        apply_workbook_updates(
            bundle.transfer,
            updates=[SheetUpdate("inv", warehouse_rows)],
            dry_run=dry_run,
        )
    )
    sheet_metrics.extend(
        apply_workbook_updates(
            bundle.perfect_order,
            updates=[SheetUpdate("new.shortages", sheet2_rows), SheetUpdate("inv", warehouse_rows)],
            dry_run=dry_run,
        )
    )
    sheet_metrics.extend(
        apply_workbook_updates(
            bundle.clr,
            updates=[SheetUpdate("Sheet1", sheet2_rows)],
            dry_run=dry_run,
        )
    )

    filters = [
        FilterSnapshot(
            name="Sales Sheet1 filter",
            input_rows=len(sales_df),
            output_rows=len(sheet1_df),
        ),
        FilterSnapshot(
            name="Sales Sheet2 filter",
            input_rows=len(sales_df),
            output_rows=len(sheet2_df),
        ),
        FilterSnapshot(
            name="BAL availability filter",
            input_rows=max(len(bal_selected_rows) - 1, 0),
            output_rows=max(len(bal_filtered_rows) - 1, 0),
        ),
        FilterSnapshot(
            name="BAL warehouse filter",
            input_rows=max(len(bal_filtered_rows) - 1, 0),
            output_rows=max(len(warehouse_rows) - 1, 0),
        ),
    ]

    duration = time.perf_counter() - start_time
    return PipelineReport(
        root=base_root,
        dry_run=dry_run,
        duration_seconds=duration,
        filters=filters,
        sheets=sheet_metrics,
    )


def trim_trailing_empty_rows(rows: List[List[Any]]) -> List[List[Any]]:
    trimmed = rows.copy()
    while trimmed and all(_is_empty(cell) for cell in trimmed[-1]):
        trimmed.pop()
    return trimmed


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def load_bal_rows(bal_path: Path) -> List[List[Any]]:
    wb = load_workbook(bal_path, data_only=True)
    try:
        ws = wb.worksheets[0]
        raw_rows = [list(row) for row in ws.iter_rows(min_col=1, max_col=8, values_only=True)]
        rows = trim_trailing_empty_rows(raw_rows)
        logging.info("BAL '%s' -> captured %d rows from sheet '%s'", bal_path.name, len(rows), ws.title)
        return rows
    finally:
        wb.close()


def _normalize_header_key(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip().casefold()


def _build_header_index(header_row: Sequence[Any]) -> Dict[str, int]:
    index: Dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        key = _normalize_header_key(raw)
        if key and key not in index:
            index[key] = idx
    return index


def select_bal_columns(rows: List[List[Any]], required_headers: Sequence[str]) -> List[List[Any]]:
    if not rows:
        return rows
    header_row = rows[0]
    header_index = _build_header_index(header_row)
    selected_indices: List[int] = []
    selected_headers: List[Any] = []
    for header in required_headers:
        key = header.strip().casefold()
        if key not in header_index:
            raise KeyError(f"Column '{header}' not found in BAL sheet header.")
        idx = header_index[key]
        selected_indices.append(idx)
        selected_headers.append(header_row[idx])
    subset_rows: List[List[Any]] = [selected_headers]
    for row in rows[1:]:
        subset_rows.append([row[idx] if idx < len(row) else None for idx in selected_indices])
    return trim_trailing_empty_rows(subset_rows)


def _is_positive_available_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return False
        return value > 0
    text = str(value).strip()
    if not text:
        return False
    if text.casefold() in BAL_NULLISH_TEXT:
        return False
    text_numeric = text.replace(",", "")
    try:
        return float(text_numeric) > 0
    except ValueError:
        return False


def filter_bal_rows_by_available(rows: List[List[Any]], header_name: str) -> List[List[Any]]:
    if not rows:
        return rows
    header_row = rows[0]
    header_index = _build_header_index(header_row)
    key = header_name.strip().casefold()
    if key not in header_index:
        raise KeyError(f"Column '{header_name}' not found in BAL sheet header.")
    idx = header_index[key]
    filtered: List[List[Any]] = [header_row]
    for row in rows[1:]:
        if idx >= len(row):
            continue
        if _is_positive_available_value(row[idx]):
            filtered.append(row)
    if len(filtered) == 1:
        return [header_row]
    return trim_trailing_empty_rows(filtered)


def detect_warehouse_column_index(rows: List[List[Any]], override_letter: Optional[str]) -> int:
    if not rows:
        raise ValueError("BAL sheet is empty; cannot detect warehouse column.")
    if override_letter:
        idx = column_index_from_string(override_letter.upper()) - 1
        if idx < 0 or idx >= len(rows[0]):
            raise ValueError(f"Override column {override_letter} is outside the captured BAL range A:H.")
        logging.debug("Using explicit warehouse column %s (index %d)", override_letter.upper(), idx)
        return idx
    for sample_row in rows[:5]:
        for idx, cell in enumerate(sample_row):
            if isinstance(cell, str) and any(keyword in cell.casefold() for keyword in WAREHOUSE_HEADER_KEYWORDS):
                logging.debug("Detected warehouse column at index %d based on header '%s'", idx, cell)
                return idx
    raise ValueError(
        "Could not detect the warehouse column automatically; specify it with --warehouse-column."
    )


def filter_bal_by_warehouse(
    rows: List[List[Any]],
    column_idx: int,
    allowed_values: Sequence[str],
) -> List[List[Any]]:
    if not rows:
        return rows
    allowed = {value.strip().upper() for value in allowed_values}
    header, data_rows = rows[0], rows[1:]
    filtered = [row for row in data_rows if _value_in_set(row, column_idx, allowed)]
    logging.info("Warehouse filter retained %d of %d BAL data rows", len(filtered), len(data_rows))
    return [header] + filtered if filtered else [header]


def _value_in_set(row: List[Any], column_idx: int, allowed: Sequence[str]) -> bool:
    if column_idx >= len(row):
        return False
    value = row[column_idx]
    if value is None:
        return False
    if isinstance(value, float) and math.isnan(value):
        return False
    value_text = str(value).strip().upper()
    return value_text in allowed


def load_sales_dataframe(sales_path: Path) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    sheet_name = resolve_sales_sheet_name(sales_path, SALES_SOURCE_SHEET)
    raw_df = pd.read_excel(sales_path, sheet_name=sheet_name, header=None, dtype=object)
    col_start = SALES_COL_START_INDEX - 1
    col_end = SALES_COL_END_INDEX
    header_values = raw_df.iloc[SALES_HEADER_ROW_INDEX, col_start:col_end].tolist()
    data_df = raw_df.iloc[SALES_HEADER_ROW_INDEX + 1 :, col_start:col_end].copy()
    data_df.columns = ALL_SALES_LETTERS
    data_df = data_df.dropna(how="all").reset_index(drop=True)
    logging.info(
        "Sales shortages '%s' -> loaded %d data rows from sheet '%s'",
        sales_path.name,
        len(data_df),
        sheet_name,
    )
    header_map = dict(zip(ALL_SALES_LETTERS, header_values))
    return data_df, header_map


def ensure_columns(df: pd.DataFrame, required: Sequence[str]) -> None:
    missing = [col for col in required if col not in df.columns]
    if missing:
        raise KeyError(f"Missing expected columns: {', '.join(missing)}")


def normalize_series(series: pd.Series) -> pd.Series:
    return series.fillna("").astype(str).str.strip()


def build_rows(
    df: pd.DataFrame,
    column_order: Sequence[str],
    header_map: Dict[str, Any],
    fallback_message: Optional[str],
    fill_value: Optional[Any] = None,
    header_fill_value: Optional[Any] = None,
) -> List[List[Any]]:
    header_row = [
        format_header_value_with_fill(header_map.get(letter), letter, header_fill_value)
        for letter in column_order
    ]
    header_row = _deduplicate_headers(header_row)
    if df.empty:
        if fill_value is not None:
            zero_row = [fill_value] * len(column_order)
            return [header_row, zero_row]
        if fallback_message:
            message_row = [fallback_message] + [""] * (len(column_order) - 1)
            return [header_row, message_row]
        return [header_row]
    subset = df.loc[:, column_order]
    if fill_value is not None:
        clean_subset = subset.fillna(fill_value)
    else:
        clean_subset = subset.where(pd.notna(subset), None)
    data_rows = clean_subset.values.tolist()
    return [header_row] + data_rows


def format_header_value(value: Any, fallback: str) -> Any:
    return _format_header_value(value, fallback, None)


def format_header_value_with_fill(
    value: Any,
    fallback: str,
    fill_value: Optional[Any],
) -> Any:
    return _format_header_value(value, fallback, fill_value)


def _format_header_value(
    value: Any,
    fallback: str,
    fill_value: Optional[Any],
) -> Any:
    if value is None:
        return fallback if fill_value is None else fill_value
    if isinstance(value, float) and math.isnan(value):
        return fallback if fill_value is None else fill_value
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("="):
            return fallback if fill_value is None else fill_value
        if text:
            return text
        if fill_value is not None:
            return fill_value
        return fallback
    return value


def _deduplicate_headers(headers: Sequence[Any]) -> List[str]:
    counts: Dict[str, int] = {}
    adjusted: List[str] = []
    for raw in headers:
        base = str(raw) if raw is not None else ""
        if not base.strip():
            base = "Column"
        index = counts.get(base, 0)
        candidate = base if index == 0 else f"{base}_{index}"
        while candidate in counts:
            index += 1
            candidate = f"{base}_{index}"
        counts[base] = index + 1
        counts[candidate] = 1
        adjusted.append(candidate)
    return adjusted


def build_sales_sheet_one(df: pd.DataFrame) -> pd.DataFrame:
    ensure_columns(df, SHEET1_REQUIRED_COLUMNS.union(SHEET1_OUTPUT_REQUIRED_COLUMNS))
    q_series = normalize_series(df["Q"])
    b_series = normalize_series(df["B"])
    ae_positive = pd.to_numeric(df["AE"], errors="coerce").fillna(0) > 0
    mask = (~q_series.isin(FILTER1_EXCLUDE_Q)) & ae_positive & (b_series != "")
    filtered = df.loc[mask].copy()
    logging.info("Sheet 1 filter retained %d rows", len(filtered))
    ensure_columns(filtered, SHEET1_OUTPUT_REQUIRED_COLUMNS)
    return filtered


def build_sales_sheet_two(df: pd.DataFrame) -> pd.DataFrame:
    ensure_columns(df, SHEET2_REQUIRED_COLUMNS.union(SHEET2_OUTPUT_REQUIRED_COLUMNS))
    q_series = normalize_series(df["Q"])
    r_series = normalize_series(df["R"])
    b_series = normalize_series(df["B"])
    ae_positive = pd.to_numeric(df["AE"], errors="coerce").fillna(0) > 0
    mask = (
        (~r_series.isin(FILTER2_EXCLUDE_R))
        & (~q_series.isin(FILTER1_EXCLUDE_Q))
        & ae_positive
        & (b_series != "")
    )
    filtered = df.loc[mask].copy()
    logging.info("Sheet 2 filter retained %d rows", len(filtered))
    ensure_columns(filtered, SHEET2_OUTPUT_REQUIRED_COLUMNS)
    return filtered


def apply_workbook_updates(
    workbook_path: Path,
    updates: Sequence[SheetUpdate],
    dry_run: bool,
) -> List[SheetMetrics]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    wb = load_workbook(workbook_path)
    metrics: List[SheetMetrics] = []
    try:
        for instruction in updates:
            rows = instruction.rows
            preserve_header = instruction.preserve_header
            ws = _resolve_sheet(
                wb,
                instruction.sheet_name,
                instruction.create_if_missing,
                dry_run,
            )
            column_count = len(rows[0]) if rows else 0
            planned_rows = max(len(rows) - 1, 0)
            logging.info(
                "%s -> %s: preparing to write %d rows",
                workbook_path.name,
                ws.title,
                planned_rows,
            )
            if dry_run:
                metrics.append(
                    SheetMetrics(
                        workbook=workbook_path,
                        sheet=ws.title,
                        rows_written=planned_rows,
                        columns_written=column_count,
                        dry_run=True,
                    )
                )
                continue
            _clear_sheet(ws, keep_header=preserve_header)
            if preserve_header:
                data_rows = rows[1:] if len(rows) > 1 else []
                _write_rows(ws, data_rows, start_row=2)
                written_rows = len(data_rows)
                if rows and column_count and data_rows:
                    _ensure_table(
                        ws,
                        f"{_sanitize_table_name(ws.title)}_Table",
                        1,
                        len(rows),
                        column_count,
                    )
            else:
                start_row, row_count, col_count = _write_rows(ws, rows)
                written_rows = max(row_count - 1, 0)
                column_count = col_count
                if row_count and col_count:
                    _ensure_table(
                        ws,
                        f"{_sanitize_table_name(ws.title)}_Table",
                        start_row,
                        row_count,
                        col_count,
                    )
            logging.info(
                "%s -> %s: wrote %d row(s)",
                workbook_path.name,
                ws.title,
                written_rows,
            )
            metrics.append(
                SheetMetrics(
                    workbook=workbook_path,
                    sheet=ws.title,
                    rows_written=written_rows,
                    columns_written=column_count,
                    dry_run=False,
                )
            )
        if dry_run:
            logging.info("Dry run active; skipping save for %s", workbook_path.name)
        else:
            remove_external_links(wb)
            wb.save(workbook_path)
            logging.info("Saved %s", workbook_path.name)
    finally:
        wb.close()
    return metrics


class _DummySheet:
    def __init__(self, title: str):
        self.title = title


def _resolve_sheet(
    wb,
    sheet_name: Optional[str],
    create_if_missing: bool = False,
    dry_run: bool = False,
):
    if sheet_name:
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        normalized = sheet_name.strip().casefold()
        for existing in wb.sheetnames:
            if existing.strip().casefold() == normalized:
                logging.info("Using sheet '%s' (matched requested '%s')", existing, sheet_name)
                return wb[existing]
        if create_if_missing:
            if dry_run:
                logging.info("Dry run: would create sheet '%s'", sheet_name)
                return _DummySheet(sheet_name)
            ws = wb.create_sheet(title=sheet_name)
            logging.info("Created sheet '%s'", sheet_name)
            return ws
        source = getattr(wb, "path", "<workbook>")
        available = ", ".join(wb.sheetnames)
        raise KeyError(
            f"Sheet '{sheet_name}' not found in workbook '{source}'. Available sheets: {available}"
        )
    return wb.worksheets[0]


def update_sales_workbook(
    sales_path: Path,
    sheet1_rows: List[List[Any]],
    sheet2_rows: List[List[Any]],
    dry_run: bool,
) -> List[SheetMetrics]:
    wb = load_workbook(sales_path)
    metrics: List[SheetMetrics] = []
    try:
        targets = [("Sheet1", sheet1_rows), ("Sheet2", sheet2_rows)]
        for name, rows in targets:
            data_rows = max(len(rows) - 1, 0)
            column_count = len(rows[0]) if rows else 0
            if dry_run:
                logging.info(
                    "Dry run: would write %d row(s) with %d column(s) to %s!%s",
                    data_rows,
                    column_count,
                    sales_path.name,
                    name,
                )
                metrics.append(
                    SheetMetrics(
                        workbook=sales_path,
                        sheet=name,
                        rows_written=data_rows,
                        columns_written=column_count,
                        dry_run=True,
                    )
                )
                continue
            if name in wb.sheetnames:
                ws = wb[name]
            else:
                ws = wb.create_sheet(title=name)
                logging.info("Created sheet '%s' in %s", name, sales_path.name)
            _clear_sheet(ws)
            start_row, row_count, col_count = _write_rows(ws, rows)
            _ensure_table(
                ws,
                f"{_sanitize_table_name(name)}_Table",
                start_row,
                row_count,
                col_count,
            )
            written_rows = max(row_count - 1, 0)
            logging.info(
                "%s -> %s: wrote %d row(s)",
                sales_path.name,
                name,
                written_rows,
            )
            metrics.append(
                SheetMetrics(
                    workbook=sales_path,
                    sheet=ws.title,
                    rows_written=written_rows,
                    columns_written=col_count,
                    dry_run=False,
                )
            )
            # remove the extranal link refresh 
        if dry_run:
            logging.info("Dry run active; skipping save for %s", sales_path.name)
        else:
            remove_external_links(wb)
            wb.save(sales_path)
            logging.info("Saved %s (Sheet1 & Sheet2 refreshed)", sales_path.name)
    finally:
        wb.close()
    return metrics


def _clear_sheet(ws, keep_header: bool = False) -> None:
    max_row = ws.max_row or 1
    if keep_header:
        if max_row > 1:
            ws.delete_rows(2, max_row - 1)
    else:
        ws.delete_rows(1, max_row)
    if hasattr(ws, "_tables"):
        ws._tables = TableList()
    if ws.auto_filter and getattr(ws.auto_filter, "ref", None):
        ws.auto_filter.ref = None


def _write_rows(
    ws,
    rows: List[List[Any]],
    start_row: int = 1,
) -> Tuple[int, int, int]:
    if not rows:
        return start_row, 0, 0
    column_count = 0
    for r_idx, row in enumerate(rows, start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
            if c_idx > column_count:
                column_count = c_idx
    written_rows = len(rows)
    return start_row, written_rows, column_count


def _ensure_table(
    ws,
    table_name: str,
    start_row: int,
    row_count: int,
    column_count: int,
) -> None:
    if row_count <= 0 or column_count <= 0:
        return
    end_column = get_column_letter(column_count)
    end_row = start_row + row_count - 1
    ref = f"A{start_row}:{end_column}{end_row}"
    if hasattr(ws, "_tables"):
        preserved = TableList()
        for table in ws._tables:
            if table.displayName != table_name:
                preserved.add(table)
        ws._tables = preserved
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def _sanitize_table_name(name: str) -> str:
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in name)
    if cleaned and cleaned[0].isdigit():
        cleaned = f"T_{cleaned}"
    return cleaned or "Table"


def remove_external_links(wb) -> None:
    links = getattr(wb, "_external_links", None)
    if links:
        wb._external_links = []
        logging.debug("Removed %d external link(s) before saving", len(links))


def orchestrate(args: argparse.Namespace) -> PipelineReport:
    root = args.root.expanduser().resolve()
    logging.info("Using root folder: %s", root)
    bundle = collect_workbook_bundle(args, root)
    report = execute_pipeline(bundle, args.warehouse_column, args.dry_run, root)
    logging.info(
        "Pipeline finished in %.2f seconds (%s)",
        report.duration_seconds,
        "dry run" if report.dry_run else "live write",
    )
    for snapshot in report.filters:
        logging.info(
            "Filter '%s': %d -> %d rows",
            snapshot.name,
            snapshot.input_rows,
            snapshot.output_rows,
        )
    for metric in report.sheets:
        logging.info(
            "Workbook %s -> %s: %d row(s), %d column(s) [%s]",
            metric.workbook.name,
            metric.sheet,
            metric.rows_written,
            metric.columns_written,
            "dry run" if metric.dry_run else "written",
        )
    return report


def launch_gui(args: argparse.Namespace) -> None:
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox, scrolledtext, ttk
    except ImportError as exc:  # pragma: no cover - depends on local Python build
        raise SystemExit("Tkinter is required for GUI mode but is not available.") from exc

    file_fields = [
        ("BAL workbook", "bal"),
        ("Sales order shortages workbook", "sales"),
        ("Raw shortages workbook", "inventory"),
        ("PPP shortages workbook", "shortages"),
        ("Color transfer workbook", "transfer"),
        ("Perfect order workbook", "perfect_order"),
        ("CLR workbook", "clr"),
    ]
    attr_map = {
        "bal": "bal_file",
        "sales": "sales_file",
        "inventory": "inventory_file",
        "shortages": "shortages_file",
        "transfer": "transfer_file",
        "perfect_order": "perfect_order_file",
        "clr": "clr_file",
    }

    class TextHandler(logging.Handler):
        def __init__(self, widget: scrolledtext.ScrolledText):
            super().__init__()
            self.widget = widget

        def emit(self, record: logging.LogRecord) -> None:
            message = self.format(record)
            self.widget.after(0, self._append, message)

        def _append(self, message: str) -> None:
            self.widget.configure(state="normal")
            self.widget.insert("end", message + "\n")
            self.widget.see("end")
            self.widget.configure(state="disabled")

    class ProcessExcelGUI:
        def __init__(self, defaults: argparse.Namespace):
            self.defaults = defaults
            self.window = tk.Tk()
            self.window.title("Excel Tracker Updater")
            self.window.minsize(780, 520)
            self.window.columnconfigure(0, weight=1)
            self.window.rowconfigure(0, weight=1)
            self.worker: Optional[threading.Thread] = None
            self.latest_report: Optional[PipelineReport] = None

            self._filedialog = filedialog
            self._ttk = ttk
            self._scrolledtext = scrolledtext

            saved = load_preferences()
            saved_root = saved.get("root") or str(defaults.root)
            default_root_path = Path(saved_root).expanduser().resolve()
            self.vars: Dict[str, Any] = {
                "root": tk.StringVar(value=str(default_root_path)),
                "warehouse": tk.StringVar(value=saved.get("warehouse", defaults.warehouse_column or "")),
                "dry_run": tk.BooleanVar(value=bool(saved.get("dry_run", defaults.dry_run))),
                "verbose": tk.BooleanVar(value=bool(saved.get("verbose", defaults.verbose))),
            }
            for key, attr in attr_map.items():
                stored_value = saved.get(key)
                if stored_value:
                    initial = stored_value
                else:
                    attr_value = getattr(defaults, attr, None)
                    initial = str(attr_value) if attr_value else ""
                self.vars[key] = tk.StringVar(value=initial)

            self.status_var = tk.StringVar(value="Idle")
            self._build_layout(file_fields)
            self.window.protocol("WM_DELETE_WINDOW", self._on_close)

        def _build_layout(self, fields):
            ttk = self._ttk
            scrolledtext = self._scrolledtext
            frame = ttk.Frame(self.window, padding=12)
            frame.grid(row=0, column=0, sticky="nsew")
            frame.columnconfigure(1, weight=1)

            row = 0
            ttk.Label(frame, text="Root folder").grid(row=row, column=0, sticky="w", pady=2)
            ttk.Entry(frame, textvariable=self.vars["root"]).grid(row=row, column=1, sticky="ew", pady=2)
            ttk.Button(frame, text="Browse", command=self._browse_root).grid(row=row, column=2, padx=(6, 0), pady=2)

            row += 1
            ttk.Label(frame, text="Warehouse column (optional)").grid(row=row, column=0, sticky="w", pady=2)
            ttk.Entry(frame, textvariable=self.vars["warehouse"]).grid(row=row, column=1, sticky="w", pady=2)

            row += 1
            ttk.Label(frame, text="Workbooks").grid(row=row, column=0, sticky="w", pady=(12, 2))
            ttk.Button(
                frame,
                text="Auto detect (fill blanks)",
                command=self._auto_detect,
            ).grid(row=row, column=2, sticky="e", pady=(12, 2))
            row += 1

            for label_text, key in fields:
                ttk.Label(frame, text=label_text).grid(row=row, column=0, sticky="w", pady=2)
                ttk.Entry(frame, textvariable=self.vars[key]).grid(row=row, column=1, sticky="ew", pady=2)
                ttk.Button(
                    frame,
                    text="Browse",
                    command=lambda k=key: self._browse_file(k),
                ).grid(row=row, column=2, padx=(6, 0), pady=2)
                row += 1

            ttk.Separator(frame).grid(row=row, column=0, columnspan=3, sticky="ew", pady=(10, 10))
            row += 1

            ttk.Checkbutton(frame, text="Dry run (no writing)", variable=self.vars["dry_run"]).grid(
                row=row,
                column=0,
                sticky="w",
            )
            ttk.Checkbutton(frame, text="Verbose logging", variable=self.vars["verbose"]).grid(
                row=row,
                column=1,
                sticky="w",
            )
            row += 1

            self.run_button = ttk.Button(frame, text="Run", command=self._start_processing)
            self.run_button.grid(row=row, column=0, sticky="w", pady=(10, 6))
            ttk.Label(frame, textvariable=self.status_var).grid(row=row, column=1, sticky="w", pady=(10, 6))
            row += 1

            ttk.Label(frame, text="Log & Reports").grid(row=row, column=0, sticky="w")
            row += 1

            self.notebook = ttk.Notebook(frame)
            self.notebook.grid(row=row, column=0, columnspan=3, sticky="nsew")
            frame.rowconfigure(row, weight=1)

            log_tab = ttk.Frame(self.notebook)
            report_tab = ttk.Frame(self.notebook)
            self.notebook.add(log_tab, text="Log")
            self.notebook.add(report_tab, text="Reports")
            self.report_tab = report_tab

            self.log_widget = scrolledtext.ScrolledText(log_tab, height=12, state="disabled", wrap="word")
            self.log_widget.pack(fill="both", expand=True)

            self.report_summary_var = tk.StringVar(value="Run the process to generate a report.")
            ttk.Label(
                report_tab,
                textvariable=self.report_summary_var,
                wraplength=520,
                justify="left",
            ).pack(anchor="w", padx=12, pady=(12, 8))

            filter_frame = ttk.LabelFrame(report_tab, text="Filter Summary")
            filter_frame.pack(fill="x", padx=12, pady=(0, 8))
            self.filter_tree = ttk.Treeview(
                filter_frame,
                columns=("stage", "input", "output"),
                show="headings",
                height=4,
            )
            self.filter_tree.heading("stage", text="Stage")
            self.filter_tree.heading("input", text="Input rows")
            self.filter_tree.heading("output", text="Output rows")
            self.filter_tree.column("stage", anchor="w", width=260)
            self.filter_tree.column("input", anchor="center", width=120)
            self.filter_tree.column("output", anchor="center", width=120)
            self.filter_tree.pack(fill="x", expand=False, padx=10, pady=8)

            work_frame = ttk.LabelFrame(report_tab, text="Workbook Updates")
            work_frame.pack(fill="both", expand=True, padx=12, pady=(0, 12))
            self.work_tree = ttk.Treeview(
                work_frame,
                columns=("workbook", "sheet", "rows", "cols", "mode"),
                show="headings",
                height=6,
            )
            self.work_tree.heading("workbook", text="Workbook")
            self.work_tree.heading("sheet", text="Sheet")
            self.work_tree.heading("rows", text="Rows")
            self.work_tree.heading("cols", text="Columns")
            self.work_tree.heading("mode", text="Mode")
            self.work_tree.column("workbook", anchor="w", width=180)
            self.work_tree.column("sheet", anchor="w", width=150)
            self.work_tree.column("rows", anchor="center", width=80)
            self.work_tree.column("cols", anchor="center", width=90)
            self.work_tree.column("mode", anchor="center", width=100)
            work_scroll = ttk.Scrollbar(work_frame, orient="vertical", command=self.work_tree.yview)
            self.work_tree.configure(yscrollcommand=work_scroll.set)
            self.work_tree.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=8)
            work_scroll.pack(side="right", fill="y", padx=(0, 8), pady=8)

        def _browse_root(self) -> None:
            directory = self._filedialog.askdirectory(
                initialdir=self.vars["root"].get() or None,
                title="Select root folder",
            )
            if directory:
                self.vars["root"].set(directory)

        def _browse_file(self, key: str) -> None:
            path = self._filedialog.askopenfilename(
                initialdir=self.vars["root"].get() or None,
                title="Select workbook",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            )
            if path:
                self.vars[key].set(path)

        def _build_args(self) -> argparse.Namespace:
            root_text = self.vars["root"].get().strip()
            base_root = Path(root_text or str(default_root())).expanduser().resolve()

            def optional_path(key: str) -> Optional[Path]:
                text = self.vars[key].get().strip()
                if not text:
                    return None
                candidate = Path(text).expanduser()
                if not candidate.is_absolute():
                    candidate = (base_root / candidate).resolve()
                return candidate

            return argparse.Namespace(
                root=base_root,
                bal_pattern=BAL_PATTERN_DEFAULT,
                sales_pattern=SALES_PATTERN_DEFAULT,
                warehouse_column=(self.vars["warehouse"].get().strip() or None),
                bal_file=optional_path("bal"),
                sales_file=optional_path("sales"),
                inventory_file=optional_path("inventory"),
                shortages_file=optional_path("shortages"),
                transfer_file=optional_path("transfer"),
                perfect_order_file=optional_path("perfect_order"),
                clr_file=optional_path("clr"), # in sheet 1 ${sheet 1} in the for sheet
                dry_run=bool(self.vars["dry_run"].get()),
                verbose=bool(self.vars["verbose"].get()),
                gui=True,
                cli=False,
            )

        def _configure_logging(self, verbose: bool) -> None:
            logger = logging.getLogger()
            level = logging.DEBUG if verbose else logging.INFO
            logger.setLevel(level)
            for handler in list(logger.handlers):
                logger.removeHandler(handler)
            gui_handler = TextHandler(self.log_widget)
            gui_handler.setFormatter(logging.Formatter("%(levelname)s: %(message)s"))
            logger.addHandler(gui_handler)

        def _clear_log(self) -> None:
            self.log_widget.configure(state="normal")
            self.log_widget.delete("1.0", "end")
            self.log_widget.configure(state="disabled")

        def _set_running(self, running: bool) -> None:
            if running:
                self.status_var.set("Running...")
                self.run_button.configure(state="disabled")
            else:
                self.status_var.set("Idle")
                self.run_button.configure(state="normal")

        def _show_dialog(self, dialog_callable, title: str, message: str) -> None:
            self._set_running(False)
            self.status_var.set("Idle")
            self._reset_report_view("No report generated.")
            dialog_callable(title, message)

        def _reset_report_view(self, status: str = "Run the process to generate a report.") -> None:
            if hasattr(self, "report_summary_var"):
                self.report_summary_var.set(status)
            for tree in (
                getattr(self, "filter_tree", None),
                getattr(self, "work_tree", None),
            ):
                if tree is None:
                    continue
                for item in tree.get_children():
                    tree.delete(item)

        def _populate_report_view(self, report: PipelineReport) -> None:
            total_rows = sum(metric.rows_written for metric in report.sheets)
            mode_text = "Dry run (no files changed)" if report.dry_run else "Live write (files updated)"
            summary_lines = [
                f"Mode: {mode_text}",
                f"Elapsed: {report.duration_seconds:.2f} s",
                f"Rows touched: {total_rows}",
                f"Root: {report.root}",
            ]
            self.report_summary_var.set("\n".join(summary_lines))

            for item in self.filter_tree.get_children():
                self.filter_tree.delete(item)
            for snapshot in report.filters:
                self.filter_tree.insert(
                    "",
                    "end",
                    values=(snapshot.name, snapshot.input_rows, snapshot.output_rows),
                )

            for item in self.work_tree.get_children():
                self.work_tree.delete(item)
            for metric in report.sheets:
                mode_label = "Dry run" if metric.dry_run else "Written"
                self.work_tree.insert(
                    "",
                    "end",
                    values=(
                        metric.workbook.name,
                        metric.sheet,
                        metric.rows_written,
                        metric.columns_written,
                        mode_label,
                    ),
                )
            self.notebook.select(self.report_tab)

        def _handle_success(self, report: PipelineReport, message: str) -> None:
            self._set_running(False)
            self.status_var.set(message)
            self.latest_report = report
            self._populate_report_view(report)
            messagebox.showinfo("Success", message)

        def _collect_preferences(self) -> Dict[str, str]:
            values: Dict[str, str] = {
                "root": self.vars["root"].get().strip(),
                "warehouse": self.vars["warehouse"].get().strip(),
                "dry_run": "True" if self.vars["dry_run"].get() else "",
                "verbose": "True" if self.vars["verbose"].get() else "",
            }
            for key in attr_map:
                values[key] = self.vars[key].get().strip()
            return values

        def _auto_detect(self) -> None:
            root_text = self.vars["root"].get().strip()
            if not root_text:
                messagebox.showerror("Auto detect failed", "Please select a root folder first.")
                return
            try:
                root_path = Path(root_text).expanduser().resolve()
            except Exception as exc:
                messagebox.showerror("Auto detect failed", f"Invalid root folder: {exc}")
                return

            try:
                detected = auto_detect_paths(root_path)
            except Exception as exc:
                messagebox.showerror("Auto detect failed", str(exc))
                return

            for key, path in detected.items():
                if not self.vars[key].get().strip():
                    self.vars[key].set(str(path))
            messagebox.showinfo("Auto detect", "Detected workbooks have been filled where fields were blank.")

        def _start_processing(self) -> None:
            try:
                args = self._build_args()
            except Exception as exc:
                messagebox.showerror("Invalid input", f"Could not read inputs: {exc}")
                return

            save_preferences(self._collect_preferences())
            self._clear_log()
            self._reset_report_view("Processing...")
            self._configure_logging(args.verbose)
            self._set_running(True)

            def worker() -> None:
                try:
                    logging.info("Starting processing...")
                    report = orchestrate(args)
                except Exception as exc:
                    logging.error("Processing failed: %s", exc)
                    self.window.after(
                        0,
                        lambda: self._show_dialog(messagebox.showerror, "Processing failed", str(exc)),
                    )
                else:
                    completion_msg = "Dry run completed." if args.dry_run else "Processing complete."
                    logging.info(completion_msg)
                    self.window.after(
                        0,
                        lambda: self._handle_success(report, completion_msg),
                    )
                finally:
                    self.worker = None

            self.worker = threading.Thread(target=worker, daemon=True)
            self.worker.start()

        def _on_close(self) -> None:
            if self.worker and self.worker.is_alive():
                if not messagebox.askyesno("Exit", "Processing is still running. Exit anyway?"):
                    return
            self.window.destroy()

        def run(self) -> None:
            self.window.mainloop()

    ProcessExcelGUI(args).run()


def main() -> None:
    args = parse_args()
    if args.gui or (len(sys.argv) == 1 and not args.cli):
        launch_gui(args)
        return
    setup_logging(args.verbose)
    try:
        orchestrate(args)
    except Exception as exc:  # pragma: no cover - surface clear error messages
        logging.error("Processing failed: %s", exc)
        raise


if __name__ == "__main__":
    main()
